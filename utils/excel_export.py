"""Genera reportes Excel y los devuelve como bytes para enviar por WhatsApp."""

import io
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


_VERDE = "1E7A3E"
_VERDE_CLARO = "D6EAD8"
_BLANCO = "FFFFFF"


def _header_row(ws, cols: list[str], row: int = 1):
    for i, col in enumerate(cols, 1):
        cell = ws.cell(row=row, column=i, value=col)
        cell.font = Font(bold=True, color=_BLANCO)
        cell.fill = PatternFill("solid", fgColor=_VERDE)
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(i)].width = max(len(col) + 4, 12)


def reporte_lluvias(registros: list[dict], anio: int = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Lluvias"
    anio = anio or date.today().year
    ws["A1"] = f"Registro de Lluvias — {anio}"
    ws["A1"].font = Font(bold=True, size=14)

    cols = ["Fecha", "mm", "Observaciones", "Registrado por"]
    _header_row(ws, cols, row=3)

    for i, r in enumerate(registros, 4):
        ws.cell(row=i, column=1, value=r.get("fecha"))
        ws.cell(row=i, column=2, value=r.get("mm"))
        ws.cell(row=i, column=3, value=r.get("observaciones", ""))
        ws.cell(row=i, column=4, value=r.get("usuario_id", ""))
        if i % 2 == 0:
            for c in range(1, 5):
                ws.cell(row=i, column=c).fill = PatternFill("solid", fgColor=_VERDE_CLARO)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def reporte_economia(registros: list[dict], periodo: str = "") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Economía"
    ws["A1"] = f"Reporte Económico {periodo}"
    ws["A1"].font = Font(bold=True, size=14)

    cols = ["Fecha", "Tipo", "Categoría", "Concepto", "Monto $", "USD", "Lote", "Chacra"]
    _header_row(ws, cols, row=3)

    total_ingresos = 0.0
    total_egresos = 0.0

    for i, r in enumerate(registros, 4):
        tipo = r.get("tipo", "")
        monto = r.get("monto", 0)
        ws.cell(row=i, column=1, value=r.get("fecha"))
        ws.cell(row=i, column=2, value=tipo.upper())
        ws.cell(row=i, column=3, value=r.get("categoria", ""))
        ws.cell(row=i, column=4, value=r.get("concepto", ""))
        ws.cell(row=i, column=5, value=monto)
        ws.cell(row=i, column=6, value=r.get("precio_usd"))
        ws.cell(row=i, column=7, value=r.get("lote_id", ""))
        ws.cell(row=i, column=8, value=r.get("chacra_id", ""))
        if tipo == "ingreso":
            total_ingresos += monto
        else:
            total_egresos += monto

    fila_total = len(registros) + 5
    ws.cell(row=fila_total, column=4, value="TOTAL INGRESOS").font = Font(bold=True)
    ws.cell(row=fila_total, column=5, value=total_ingresos).font = Font(bold=True)
    ws.cell(row=fila_total + 1, column=4, value="TOTAL EGRESOS").font = Font(bold=True)
    ws.cell(row=fila_total + 1, column=5, value=total_egresos).font = Font(bold=True)
    ws.cell(row=fila_total + 2, column=4, value="RESULTADO NETO").font = Font(bold=True, color=_VERDE)
    ws.cell(row=fila_total + 2, column=5, value=total_ingresos - total_egresos).font = Font(bold=True, color=_VERDE)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def reporte_hacienda(lotes: list[dict], potreros: list[dict]) -> bytes:
    wb = Workbook()

    # Hoja Lotes
    ws1 = wb.active
    ws1.title = "Lotes"
    ws1["A1"] = "Estado de Lotes"
    ws1["A1"].font = Font(bold=True, size=14)
    cols_lotes = ["Lote", "Categoría", "Potrero", "Fecha Ingreso", "Origen", "Activo"]
    _header_row(ws1, cols_lotes, row=3)
    for i, l in enumerate(lotes, 4):
        ws1.cell(row=i, column=1, value=l.get("nombre"))
        ws1.cell(row=i, column=2, value=l.get("categoria"))
        ws1.cell(row=i, column=3, value=l.get("potrero_id"))
        ws1.cell(row=i, column=4, value=l.get("fecha_ingreso"))
        ws1.cell(row=i, column=5, value=l.get("origen"))
        ws1.cell(row=i, column=6, value="Sí" if l.get("activo") else "No")

    # Hoja Potreros
    ws2 = wb.create_sheet("Potreros")
    ws2["A1"] = "Estado de Potreros"
    ws2["A1"].font = Font(bold=True, size=14)
    cols_pot = ["Nombre", "Superficie (has)", "Estado"]
    _header_row(ws2, cols_pot, row=3)
    for i, p in enumerate(potreros, 4):
        ws2.cell(row=i, column=1, value=p.get("nombre"))
        ws2.cell(row=i, column=2, value=p.get("superficie_has"))
        ws2.cell(row=i, column=3, value=p.get("estado"))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
